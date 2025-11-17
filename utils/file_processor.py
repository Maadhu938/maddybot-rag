"""Utility functions for processing various file types."""
import os
import base64
from typing import Optional, Dict
from pathlib import Path

try:
    from PIL import Image
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

try:
    import PyPDF2
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from docx import Document
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def extract_text_from_file(file_path: str, file_type: str) -> Dict[str, any]:
    """Extract text content from various file types."""
    result = {
        "success": False,
        "content": "",
        "metadata": {},
        "error": None
    }
    
    try:
        file_ext = Path(file_path).suffix.lower()
        
        # Text files
        if file_ext in ['.txt', '.md', '.py', '.js', '.html', '.css', '.json', '.xml', '.csv', '.log']:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                result["content"] = f.read()
                result["success"] = True
                result["metadata"] = {"type": "text", "encoding": "utf-8"}
        
        # PDF files
        elif file_ext == '.pdf' and PDF_AVAILABLE:
            text_content = []
            with open(file_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                result["metadata"]["pages"] = len(pdf_reader.pages)
                for page_num, page in enumerate(pdf_reader.pages):
                    text_content.append(f"--- Page {page_num + 1} ---\n{page.extract_text()}")
                result["content"] = "\n\n".join(text_content)
                result["success"] = True
                result["metadata"]["type"] = "pdf"
        
        # Word documents
        elif file_ext in ['.docx', '.doc'] and DOCX_AVAILABLE:
            doc = Document(file_path)
            paragraphs = [para.text for para in doc.paragraphs]
            result["content"] = "\n".join(paragraphs)
            result["success"] = True
            result["metadata"] = {"type": "docx", "paragraphs": len(paragraphs)}
        
        # Excel files
        elif file_ext in ['.xlsx', '.xls'] and EXCEL_AVAILABLE:
            workbook = openpyxl.load_workbook(file_path)
            content_parts = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                content_parts.append(f"--- Sheet: {sheet_name} ---")
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                    content_parts.append(row_text)
            result["content"] = "\n".join(content_parts)
            result["success"] = True
            result["metadata"] = {"type": "excel", "sheets": workbook.sheetnames}
        
        else:
            result["error"] = f"Unsupported file type: {file_ext}"
            result["content"] = f"[File type {file_ext} is not yet supported for text extraction]"
            result["success"] = False
    
    except Exception as e:
        result["error"] = str(e)
        result["success"] = False
    
    return result


def process_image(image_path: str) -> Dict[str, any]:
    """Process image file and return base64 encoded data."""
    result = {
        "success": False,
        "base64": None,
        "metadata": {},
        "error": None
    }
    
    if not PIL_AVAILABLE:
        result["error"] = "PIL/Pillow not available"
        return result
    
    try:
        with Image.open(image_path) as img:
            # Get image metadata
            result["metadata"] = {
                "format": img.format,
                "mode": img.mode,
                "size": img.size,
                "width": img.width,
                "height": img.height
            }
            
            # Convert to RGB if necessary (for JPEG compatibility)
            if img.mode != 'RGB':
                img = img.convert('RGB')
            
            # Resize if too large (max 2048px on longest side)
            max_size = 2048
            if max(img.size) > max_size:
                ratio = max_size / max(img.size)
                new_size = (int(img.width * ratio), int(img.height * ratio))
                img = img.resize(new_size, Image.Resampling.LANCZOS)
                result["metadata"]["resized"] = True
                result["metadata"]["original_size"] = img.size
            
            # Save to bytes and encode to base64
            import io
            buffer = io.BytesIO()
            img.save(buffer, format='JPEG', quality=85)
            buffer.seek(0)
            result["base64"] = base64.b64encode(buffer.read()).decode('utf-8')
            result["success"] = True
    
    except Exception as e:
        result["error"] = str(e)
        result["success"] = False
    
    return result


def get_file_info(file_path: str) -> Dict[str, any]:
    """Get basic file information."""
    try:
        stat = os.stat(file_path)
        return {
            "name": os.path.basename(file_path),
            "size": stat.st_size,
            "size_mb": round(stat.st_size / (1024 * 1024), 2),
            "extension": Path(file_path).suffix.lower(),
            "exists": True
        }
    except Exception as e:
        return {
            "name": os.path.basename(file_path),
            "exists": False,
            "error": str(e)
        }

